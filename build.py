#!/usr/bin/env python3
"""Build index.html: extract data from asta.xlsx + listone.xlsx, merge research.json,
inject everything into app_template.html as the DATA constant."""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent


def norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9 ]", "", s.lower()).strip()


def extract_listone():
    wb = openpyxl.load_workbook(HERE / "listone.xlsx", data_only=True)
    players = []
    for r in wb["Tutti"].iter_rows(min_row=3, values_only=True):
        if r[3] is None:
            continue
        players.append({
            "id": r[0], "r": r[1], "rm": r[2], "name": r[3],
            "team": r[4], "qta": r[5], "fvm": r[11],
        })
    return players


def extract_asta():
    wb = openpyxl.load_workbook(HERE / "asta.xlsx", data_only=True)
    slots = []
    for row in wb["Asta"].iter_rows(min_row=17, max_row=41, values_only=True):
        if not row[0]:
            continue
        slots.append({
            "id": row[0], "role": row[1], "profile": row[2], "target": row[3],
            "team": row[4], "qta": row[5], "budget": row[6],
        })
    wishlist = {}
    sheet_role = {"Portieri": "P", "Difensori": "D", "Centrocampisti": "C", "Attaccanti": "A"}
    for sheet, role in sheet_role.items():
        ws = wb[sheet]
        groups = []
        cur = None
        for row in ws.iter_rows(values_only=True):
            a = row[0]
            if isinstance(a, str) and a.startswith("SLOT"):
                m = re.match(r"SLOT (\S+) — (.+?)\s+·\s+budget consigliato: (\d+)", a)
                cur = {
                    "slot": m.group(1) if m else a.split()[1],
                    "profile": m.group(2) if m else a,
                    "budget": int(m.group(3)) if m else None,
                    "players": [],
                }
                groups.append(cur)
            elif cur is not None and isinstance(a, (int, float)) and row[1]:
                cur["players"].append({
                    "prio": int(a), "name": row[1], "team": row[2], "qta": row[3],
                    "fvm": row[4], "tit": row[5], "rig": row[6], "bonus": row[7],
                    "maxprice": row[8], "note": row[9],
                })
        wishlist[role] = groups
    return slots, wishlist


def apply_ops(wishlist, ops):
    """ops: list of {op, role, slot, name, ...}. Supported ops:
    remove          — drop player `name` from slot list
    add             — append/insert player dict at `prio` (others reflow)
    update          — merge given fields into matching player
    """
    for o in ops:
        groups = wishlist.get(o["role"], [])
        grp = next((g for g in groups if g["slot"] == o["slot"]), None)
        if grp is None:
            print(f"  !! op skipped, slot not found: {o}")
            continue
        if o["op"] == "remove":
            before = len(grp["players"])
            grp["players"] = [p for p in grp["players"] if norm(p["name"]) != norm(o["name"])]
            if len(grp["players"]) == before:
                print(f"  !! remove: name not found: {o}")
        elif o["op"] == "update":
            hit = False
            for p in grp["players"]:
                if norm(p["name"]) == norm(o["name"]):
                    p.update({k: v for k, v in o.items() if k not in ("op", "role", "slot")})
                    hit = True
            if not hit:
                print(f"  !! update: name not found: {o}")
        elif o["op"] == "add":
            entry = {k: v for k, v in o.items() if k not in ("op", "role", "slot")}
            entry.setdefault("prio", len(grp["players"]) + 1)
            grp["players"].append(entry)
        # reflow priorities by 'prio' then original order
        grp["players"].sort(key=lambda p: p.get("prio", 99))
        for i, p in enumerate(grp["players"], 1):
            p["prio"] = i


def compute_scores(players, wishlist, research):
    """FantaScore 0-99: FVM percentile in role + titolarità + rigori/piazzati
    − infortuni − cartellini − rischio mercato. Stored as p['score']."""
    import bisect

    wentries = {}
    for groups in wishlist.values():
        for g in groups:
            for w in g["players"]:
                wentries.setdefault(norm(w["name"]), []).append(w)
    alerts = {}
    for a in research.get("alerts", []):
        alerts.setdefault((norm(a["name"]), norm(a.get("team", ""))), []).append(a)
    rigoristi = research.get("rigoristi", {})

    fvm_by_role = {}
    for p in players:
        fvm_by_role.setdefault(p["r"], []).append(p["fvm"] or 0)
    for r in fvm_by_role:
        fvm_by_role[r].sort()

    LONG_OUT = ("crociato", "operat", "mesi", "ottobre", "novembre", "dicembre",
                "settembre", "salta le prime", "fino a")
    GONE_MKT = ("lasciarlo", "depenn", "via dall", "fuori dal campionato",
                "non comprar", "fuori rosa")

    for p in players:
        vals = fvm_by_role[p["r"]]
        pct = bisect.bisect_left(vals, p["fvm"] or 0) / max(1, len(vals) - 1)
        s = 55 * pct

        ws = wentries.get(norm(p["name"]), [])
        tit = " ".join((w.get("tit") or "") for w in ws).lower()
        rig = " ".join((w.get("rig") or "") for w in ws).lower()
        if any(k in tit for k in ("altissima", "titolare", "inamovibile", "favorito", "promosso")):
            s += 20
        elif any(k in tit for k in ("uscita", "panchina", "riserva", "fondo", "out", "evitare", "al psg", "conteso a 3")):
            s += 0
        elif "alta" in tit:
            s += 14
        elif any(k in tit for k in ("ballottaggio", "conteso", "50", "monitorare", "dubbio", "scommessa")):
            s += 6
        else:
            s += 8  # nessuna informazione: neutro

        if rig.startswith(("sì", "si ")) or any(k in rig for k in ("1°", "pole", "blindato", "eredita")):
            s += 12
        elif any(k in rig for k in ("conteso", "in corsa", "alternanza", "2/2")):
            s += 6

        team_rig = rigoristi.get(p["team"], {})
        placed = norm(team_rig.get("punizioni", "") + " " + team_rig.get("corner", ""))
        pname = norm(p["name"])
        first = pname.split(" ")[0] if pname else ""
        if pname and (pname in placed or (len(first) > 3 and first in placed)):
            s += 5

        for a in alerts.get((pname, norm(p["team"])), []):
            t, txt = a.get("type", ""), (a.get("text") or "").lower()
            if t == "infortunio":
                s -= 20 if any(k in txt for k in LONG_OUT) else 10
            elif t == "fairplay":
                s -= 8
            elif t == "mercato":
                s -= 15 if any(k in txt for k in GONE_MKT) else 4
            elif t == "squalifica":
                s -= 3
            elif t == "ballottaggio":
                s -= 4
            elif t == "ok":
                s += 3
        p["score"] = max(1, min(99, round(s)))


def main():
    players = extract_listone()
    slots, wishlist = extract_asta()
    research = {}
    rpath = HERE / "research.json"
    if rpath.exists():
        research = json.loads(rpath.read_text())
        apply_ops(wishlist, research.get("wishlist_ops", []))
    compute_scores(players, wishlist, research)

    data = {
        "season": "2026-27",
        "budget": 1000,
        "roleBudget": {"P": 80, "D": 130, "C": 265, "A": 525},
        "roleSlots": {"P": 3, "D": 8, "C": 8, "A": 6},
        "players": players,
        "slots": slots,
        "wishlist": wishlist,
        "rigoristi": research.get("rigoristi", {}),
        "alerts": research.get("alerts", []),
        "sleepers": research.get("sleepers", {}),
        "researchDate": research.get("date", ""),
        "sources": research.get("sources", []),
    }

    tpl = (HERE / "app_template.html").read_text()
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    html = tpl.replace("/*__DATA__*/null", payload)
    (HERE / "index.html").write_text(html)
    print(f"index.html written: {len(html)//1024} KB, {len(players)} players, "
          f"{len(data['alerts'])} alerts, {sum(len(v) for v in data['sleepers'].values()) if data['sleepers'] else 0} sleepers")


if __name__ == "__main__":
    main()
